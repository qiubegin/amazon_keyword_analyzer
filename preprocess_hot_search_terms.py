#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Preprocess Amazon hot search term exports.

The script scans the source folder for new .csv files, removes rows whose
search term contains configured stopwords, keeps the required columns, adds
top-3 ASIN share totals, and writes processed .csv files.

It supports:
- real CSV files
- .xlsx files
- .xlsx files that were renamed to .csv
"""

import csv
import argparse
import json
import os
import re
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl. Please install openpyxl first.") from exc


BASE_DIR = Path(__file__).resolve().parent
SOURCE_ROOT = BASE_DIR / "raw_data"
OUTPUT_ROOT = BASE_DIR / "processed_data"
FILTER_RULES_DIR = BASE_DIR / "filter_rules"
STOPWORDS_FILE_PATH = FILTER_RULES_DIR / "知名品牌名筛选_停用词.xlsx"
STOPWORDS_SHEET_NAME = "筛选结果"
EXCLUDED_CATEGORIES_FILE_PATH = FILTER_RULES_DIR / "剔除类目DE.csv"
DEFAULT_SITE = "DE"
SITES = ["DE", "FR", "IT", "ES", "AU"]

REQUIRED_COLUMNS = [
    "搜索词",
    "搜索频率排名",
    "点击量最高的类别 #1",
    "点击量最高的类别 #2",
    "点击量最高的类别 #3",
    "点击量最高的商品 #1：ASIN",
    "点击量最高的商品 #2：ASIN",
    "点击量最高的商品 #3：ASIN",
    "点击量最高的商品 #1：点击份额",
    "点击量最高的商品 #2：点击份额",
    "点击量最高的商品 #3：点击份额",
    "点击量最高的商品 #1：转化份额",
    "点击量最高的商品 #2：转化份额",
    "点击量最高的商品 #3：转化份额",
    "报告日期",
]

CLICK_SHARE_COLUMNS = [
    "点击量最高的商品 #1：点击份额",
    "点击量最高的商品 #2：点击份额",
    "点击量最高的商品 #3：点击份额",
]

CONVERSION_SHARE_COLUMNS = [
    "点击量最高的商品 #1：转化份额",
    "点击量最高的商品 #2：转化份额",
    "点击量最高的商品 #3：转化份额",
]

CATEGORY_COLUMNS = [
    "点击量最高的类别 #1",
    "点击量最高的类别 #2",
    "点击量最高的类别 #3",
]

OUTPUT_COLUMNS = REQUIRED_COLUMNS + [
    "前三ASIN点击份额占比",
    "前三ASIN转化份额占比",
]

COLUMN_ALIASES = {
    "搜索词": ["搜索词"],
    "搜索频率排名": ["搜索频率排名"],
    "点击量最高的类别 #1": ["点击量最高的类别 #1", "点击量最高的分类"],
    "点击量最高的类别 #2": ["点击量最高的类别 #2", "点击量第二的分类"],
    "点击量最高的类别 #3": ["点击量最高的类别 #3", "点击量第 3 的分类", "点击量第3的分类"],
    "点击量最高的商品 #1：ASIN": ["点击量最高的商品 #1：ASIN", "点击量第1的商品：ASIN"],
    "点击量最高的商品 #2：ASIN": ["点击量最高的商品 #2：ASIN", "点击量第2的商品：ASIN"],
    "点击量最高的商品 #3：ASIN": ["点击量最高的商品 #3：ASIN", "点击量第三的商品：ASIN"],
    "点击量最高的商品 #1：点击份额": ["点击量最高的商品 #1：点击份额", "点击量最高的商品：点击份额"],
    "点击量最高的商品 #2：点击份额": ["点击量最高的商品 #2：点击份额", "点击量第二的商品：点击份额"],
    "点击量最高的商品 #3：点击份额": ["点击量最高的商品 #3：点击份额", "点击量第三的商品：点击份额"],
    "点击量最高的商品 #1：转化份额": ["点击量最高的商品 #1：转化份额", "点击量第1的商品：转化贡献占比"],
    "点击量最高的商品 #2：转化份额": ["点击量最高的商品 #2：转化份额", "热门点击商品第2名：转化贡献占比"],
    "点击量最高的商品 #3：转化份额": ["点击量最高的商品 #3：转化份额", "点击量第3的商品：转化贡献占比"],
    "报告日期": ["报告日期"],
}


def log(message):
    print(f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {message}")


def normalize_text(text):
    text = "" if text is None else str(text).strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[®™©.,:;!?'\"/\\|_+*#()\[\]{}<>~`^=]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def load_stopwords():
    if not STOPWORDS_FILE_PATH.exists():
        log(f"[WARN] Stopwords file not found: {STOPWORDS_FILE_PATH}")
        return []

    workbook = load_workbook(STOPWORDS_FILE_PATH, read_only=True, data_only=True)
    sheet = workbook[STOPWORDS_SHEET_NAME] if STOPWORDS_SHEET_NAME in workbook.sheetnames else workbook[workbook.sheetnames[0]]

    stopwords = set()
    for row in sheet.iter_rows(values_only=True):
        for value in row:
            term = normalize_text(value)
            if term and term not in {"搜索词", "项目", "内容"}:
                stopwords.add(term)

    workbook.close()
    return sorted(stopwords, key=len, reverse=True)


def load_excluded_categories():
    if not EXCLUDED_CATEGORIES_FILE_PATH.exists():
        log(f"[WARN] Excluded categories file not found: {EXCLUDED_CATEGORIES_FILE_PATH}")
        return set()

    excluded = set()
    encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1", "cp1252"]
    for encoding in encodings:
        try:
            with EXCLUDED_CATEGORIES_FILE_PATH.open("r", encoding=encoding, errors="ignore", newline="") as f:
                reader = csv.reader(f)
                for row in reader:
                    if not row:
                        continue
                    term = str(row[0]).strip()
                    normalized = normalize_text(term)
                    if not normalized:
                        continue
                    if normalized in {"类目 de", "剔除类目 de", "category", "categories"}:
                        continue
                    # The first row is a title. Keep only category-like values.
                    if term.upper() == term or "_" in term:
                        excluded.add(normalized)
            if excluded:
                return excluded
        except UnicodeError:
            continue

    return excluded


def build_stopword_matcher(stopwords):
    token_sets = [(term, set(term.split())) for term in stopwords]

    def contains_stopword(search_term):
        normalized = normalize_text(search_term)
        if not normalized:
            return False

        tokens = set(normalized.split())
        for stopword, stopword_tokens in token_sets:
            if stopword == normalized or stopword_tokens.issubset(tokens):
                return True
        return False

    return contains_stopword


def build_category_matcher(excluded_categories):
    def contains_excluded_category(row_dict):
        for column in CATEGORY_COLUMNS:
            category = normalize_text(row_dict.get(column, ""))
            if category and category in excluded_categories:
                return True
        return False

    return contains_excluded_category


def is_xlsx_content(path):
    with open(path, "rb") as f:
        return f.read(4) == b"PK\x03\x04"


def cell_to_text(value):
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def iter_excel_rows(path):
    # Pass a binary file handle so openpyxl can read xlsx content even when
    # the file extension was manually changed to .csv.
    handle = open(path, "rb")
    workbook = load_workbook(handle, read_only=True, data_only=True)
    try:
        sheet = workbook[workbook.sheetnames[0]]
        header = None
        for row in sheet.iter_rows(values_only=True):
            values = [cell_to_text(value) for value in row]
            if header is None:
                if values and values[0] == "搜索频率排名":
                    header = values
                    yield header
                continue
            if len(values) < len(header):
                values.extend([""] * (len(header) - len(values)))
            elif len(values) > len(header):
                values = values[:len(header)]
            yield values
    finally:
        workbook.close()
        handle.close()


def sniff_csv_dialect(path, encoding):
    with open(path, "r", encoding=encoding, errors="ignore", newline="") as f:
        sample = f.read(8192)
    if ";" in sample and sample.count(";") > sample.count(","):
        return csv.excel, ";"
    return csv.excel, ","


def iter_csv_rows(path):
    encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1", "cp1252"]
    last_error = None

    for encoding in encodings:
        try:
            _, delimiter = sniff_csv_dialect(path, encoding)
            with open(path, "r", encoding=encoding, errors="ignore", newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                header = None
                for row in reader:
                    if not row or len(row) < 2:
                        continue
                    if row[0].startswith("报告范围"):
                        continue
                    if header is None:
                        if row[0] == "搜索频率排名" or "搜索频率排名" in row[0]:
                            header = row
                            yield header
                        continue
                    if len(row) < len(header):
                        row.extend([""] * (len(header) - len(row)))
                    elif len(row) > len(header):
                        row = row[:len(header)]
                    yield row
            if header is not None:
                return
        except UnicodeError as exc:
            last_error = exc
            continue

    raise RuntimeError(f"Unable to read CSV file {path}: {last_error}")


def iter_source_rows(path):
    if is_xlsx_content(path):
        yield from iter_excel_rows(path)
    else:
        yield from iter_csv_rows(path)


def parse_share(value):
    if value is None:
        return 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace("%", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def format_percent(value):
    text = f"{value:.2f}".rstrip("0").rstrip(".")
    return f"{text}%"


def normalize_site(value):
    site = str(value or DEFAULT_SITE).strip().upper()
    return site if site in SITES else DEFAULT_SITE


def site_source_dir(site):
    return SOURCE_ROOT / normalize_site(site)


def site_output_dir(site):
    return OUTPUT_ROOT / normalize_site(site)


def site_manifest_path(site):
    return site_output_dir(site) / "processed_manifest.json"


def source_file_pattern(site):
    return f"{normalize_site(site)}_*Week_*.csv"


def load_manifest(manifest_path):
    if not manifest_path.exists():
        return {}
    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_manifest(manifest, manifest_path):
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")


def source_signature(path):
    stat = path.stat()
    return {
        "size": stat.st_size,
        "mtime": stat.st_mtime,
    }


def make_output_path(source_path, output_dir):
    return output_dir / f"{source_path.stem}_processed.csv"


def should_skip_source(path, manifest, output_dir):
    key = str(path.resolve())
    output_path = make_output_path(path, output_dir)
    if not output_path.exists():
        return False
    if key not in manifest:
        return True
    return manifest[key].get("signature") == source_signature(path)


def is_empty_required_row(row_dict):
    """Drop rows where all required output fields are blank."""
    return all(not str(row_dict.get(column, "")).strip() for column in REQUIRED_COLUMNS)


def build_header_map(header):
    raw_header_map = {name: idx for idx, name in enumerate(header)}
    header_map = {}

    for output_column, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            if alias in raw_header_map:
                header_map[output_column] = raw_header_map[alias]
                break

    return header_map


def process_file(source_path, output_dir, contains_excluded_category, contains_stopword):
    output_path = make_output_path(source_path, output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    row_iter = iter_source_rows(source_path)
    try:
        header = next(row_iter)
    except StopIteration:
        raise RuntimeError("No header row found")

    header_map = build_header_map(header)
    missing_columns = [column for column in REQUIRED_COLUMNS if column not in header_map]
    if missing_columns:
        raise RuntimeError(f"Missing required columns: {missing_columns}")

    total_rows = 0
    kept_rows = 0
    removed_category_rows = 0
    removed_stopword_rows = 0
    removed_empty_rows = 0

    with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()

        for row in row_iter:
            if not row or len(row) < len(header):
                continue

            total_rows += 1
            row_dict = {column: row[header_map[column]] for column in REQUIRED_COLUMNS}
            row_dict = {column: str(value).strip() for column, value in row_dict.items()}
            search_term = row_dict["搜索词"]

            if not search_term or is_empty_required_row(row_dict):
                removed_empty_rows += 1
                continue

            if contains_excluded_category(row_dict):
                removed_category_rows += 1
                continue

            if contains_stopword(search_term):
                removed_stopword_rows += 1
                continue

            click_total = sum(parse_share(row_dict[column]) for column in CLICK_SHARE_COLUMNS)
            conversion_total = sum(parse_share(row_dict[column]) for column in CONVERSION_SHARE_COLUMNS)
            row_dict["前三ASIN点击份额占比"] = format_percent(click_total)
            row_dict["前三ASIN转化份额占比"] = format_percent(conversion_total)

            writer.writerow(row_dict)
            kept_rows += 1

            if total_rows % 50000 == 0:
                removed_total = removed_category_rows + removed_stopword_rows + removed_empty_rows
                log(f"   processed={total_rows:,}, kept={kept_rows:,}, removed={removed_total:,}")

    return {
        "output": str(output_path),
        "total_rows": total_rows,
        "kept_rows": kept_rows,
        "removed_rows": removed_category_rows + removed_stopword_rows + removed_empty_rows,
        "removed_category_rows": removed_category_rows,
        "removed_stopword_rows": removed_stopword_rows,
        "removed_empty_rows": removed_empty_rows,
    }


def find_source_files(site):
    source_dir = site_source_dir(site)
    output_dir = site_output_dir(site)
    files = []
    for path in source_dir.glob(source_file_pattern(site)):
        if path.name.startswith("~$"):
            continue
        if path.parent == output_dir:
            continue
        if path.name.endswith("_processed.csv"):
            continue
        files.append(path)
    return sorted(files)


def parse_args():
    parser = argparse.ArgumentParser(description="Preprocess Amazon hot search term CSV exports.")
    parser.add_argument(
        "--force",
        action="store_true",
        help="Reprocess files even if they are unchanged and already exist in the manifest.",
    )
    parser.add_argument(
        "--site",
        choices=["all", *SITES],
        default="all",
        help="Site code to process. Defaults to all configured site folders.",
    )
    return parser.parse_args()


def process_site(site, args, contains_excluded_category, contains_stopword):
    site = normalize_site(site)
    source_dir = site_source_dir(site)
    output_dir = site_output_dir(site)
    manifest_path = site_manifest_path(site)
    source_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    log(f"[{site}] Source folder: {source_dir}")
    log(f"[{site}] Output folder: {output_dir}")

    manifest = load_manifest(manifest_path)
    source_files = find_source_files(site)
    processed_count = 0
    skipped_count = 0
    for source_path in source_files:
        if not args.force and should_skip_source(source_path, manifest, output_dir):
            log(f"[{site}] Skip unchanged file: {source_path.name}")
            skipped_count += 1
            continue

        log(f"[{site}] Processing: {source_path.name}")
        try:
            result = process_file(source_path, output_dir, contains_excluded_category, contains_stopword)
        except (BadZipFile, RuntimeError, OSError) as exc:
            log(f"[{site}][ERROR] Failed: {source_path.name} - {exc}")
            continue

        manifest[str(source_path.resolve())] = {
            "signature": source_signature(source_path),
            "processed_at": datetime.now().isoformat(timespec="seconds"),
            **result,
        }
        save_manifest(manifest, manifest_path)
        processed_count += 1
        log(
            f"[{site}] Done: "
            f"{source_path.name}, total={result['total_rows']:,}, "
            f"kept={result['kept_rows']:,}, "
            f"removed_categories={result['removed_category_rows']:,}, "
            f"removed_stopwords={result['removed_stopword_rows']:,}, "
            f"removed_empty={result['removed_empty_rows']:,}, "
            f"output={result['output']}"
        )

    if not source_files:
        log(f"[{site}] No source .csv files found.")
    log(f"[{site}] Finished. processed={processed_count}, skipped={skipped_count}")
    return processed_count, skipped_count


def main():
    args = parse_args()
    sites = SITES if args.site == "all" else [normalize_site(args.site)]
    log(f"Sites: {', '.join(sites)}")

    excluded_categories = load_excluded_categories()
    contains_excluded_category = build_category_matcher(excluded_categories)
    log(f"Loaded excluded categories: {len(excluded_categories)}")

    stopwords = load_stopwords()
    contains_stopword = build_stopword_matcher(stopwords)
    log(f"Loaded stopwords: {len(stopwords)}")

    total_processed = 0
    total_skipped = 0
    for site in sites:
        processed_count, skipped_count = process_site(site, args, contains_excluded_category, contains_stopword)
        total_processed += processed_count
        total_skipped += skipped_count

    log(f"Finished all sites. processed={total_processed}, skipped={total_skipped}")


if __name__ == "__main__":
    main()
